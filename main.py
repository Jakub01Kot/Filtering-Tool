import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def append_to_spreadsheet(source_file, existing_file):
    # Load the data from the source file
    wb_links = load_workbook(source_file, data_only=True)
    ws_links = wb_links.active

    # Extract LinkedIn URLs from the hyperlinks in the 'Name' column
    linkedin_urls = []
    for i, row in enumerate(ws_links.iter_rows(min_row=2, max_row=ws_links.max_row, min_col=1, max_col=1)):
        hyperlink = row[0].hyperlink
        if hyperlink is not None:
            linkedin_urls.append(hyperlink.target)
        else:
            linkedin_urls.append(None)

    df_input = pd.read_excel(source_file)

    # Load the existing Excel workbook and worksheet
    wb_existing = load_workbook(existing_file)
    ws = wb_existing.active

    # Convert the existing Excel data to a dataframe for easier comparison
    data = ws.values
    columns = next(data)[0:]  # Assumes first line is header
    df_existing = pd.DataFrame(data, columns=columns)

    # Match on Name and Full Name columns to get unique rows
    merged_df = df_input.merge(df_existing, left_on="Name", right_on="Full Name", how='left', indicator=True)
    unique_rows = merged_df[merged_df['_merge'] == 'left_only']

    # Drop the merge column and any other unneeded columns
    unique_rows = unique_rows.drop(columns=['_merge', 'Name', 'Profession', 'Experience'])

    columns_order = ['Location', 'Swiss Connection', 'Source (url)',
                     'Reason for selection (criteria used)', 'Full Name',
                     'Title', 'Company', 'LinkedIn (url)']

    unique_rows = unique_rows[columns_order]

    # Find the starting row to append data in existing_file
    start_row = 82
    while ws.cell(row=start_row, column=1).value:
        start_row += 1

    # Initialize existing_urls as an empty set
    existing_urls = set()

    # Populate existing_urls with existing LinkedIn URLs from ws
    for row in data[81:]:
        existing_url = row[7]  # Assuming LinkedIn (url) is in the 8th column
        if isinstance(existing_url, str) and existing_url.startswith("http"):
            existing_urls.add(existing_url)

    for index, row in unique_rows.iterrows():
        existing_url = linkedin_urls[index]  # Get the existing URL for the current row
        if isinstance(existing_url, str) and existing_url.startswith("http"):
            if existing_url not in existing_urls:  # Use existing_urls to check for duplicates
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=start_row, column=c_idx, value=value)
                    # Set hyperlink for LinkedIn(url) column
                    if c_idx == 8 and isinstance(value, str) and value.startswith("http"):
                        original_link_formula = ws_links.cell(row=index + 2, column=1).hyperlink.target
                        ws.cell(row=start_row, column=c_idx).hyperlink = original_link_formula
                start_row += 1

    # Save the modified workbook
    wb_existing.save(existing_file)

if __name__ == "__main__":
    source_file = '/Users/jakub/Downloads/Output_with_links.xlsx'
    existing_file = '/Users/jakub/Downloads/Existing_spreadsheet.xlsx'
    append_to_spreadsheet(source_file, existing_file)
